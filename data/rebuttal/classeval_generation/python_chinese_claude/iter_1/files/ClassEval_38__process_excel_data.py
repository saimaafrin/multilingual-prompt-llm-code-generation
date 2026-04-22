class _M:
    import openpyxl
    from openpyxl import load_workbook, Workbook
    
    class ExcelProcessor:
        def __init__(self):
            pass
        
        def process_excel_data(self, N, save_file_name):
            """
            将Excel文件中指定列的内容转换为大写
            :param N: int, 要更改的列的序号
            :param save_file_name: str, 源文件名
            :return:(int, str), 前者是write_excel的返回值，而后者是处理后数据的保存文件名
            >>> processor = ExcelProcessor()
            >>> success, output_file = processor.process_excel_data(1, 'test_data.xlsx')
            """
            try:
                # 加载Excel文件
                workbook = load_workbook(save_file_name)
                sheet = workbook.active
                
                # 遍历指定列的所有单元格，将内容转换为大写
                for row in range(1, sheet.max_row + 1):
                    cell = sheet.cell(row=row, column=N)
                    if cell.value is not None and isinstance(cell.value, str):
                        cell.value = cell.value.upper()
                
                # 生成输出文件名
                if save_file_name.endswith('.xlsx'):
                    output_file_name = save_file_name[:-5] + '_processed.xlsx'
                else:
                    output_file_name = save_file_name + '_processed.xlsx'
                
                # 保存处理后的文件
                workbook.save(output_file_name)
                workbook.close()
                
                # 调用write_excel方法（假设返回1表示成功）
                result = self.write_excel(output_file_name)
                
                return (result, output_file_name)
            
            except Exception as e:
                # 如果出现错误，返回0和空字符串
                return (0, "")
        
        def write_excel(self, file_name):
            """
            辅助方法，用于验证Excel文件写入
            :param file_name: str, 文件名
            :return: int, 1表示成功，0表示失败
            """
            try:
                # 验证文件是否存在且可读
                workbook = load_workbook(file_name)
                workbook.close()
                return 1
            except:
                return 0