class _M:
    def process_excel_data(self, N, save_file_name):
        """
        Cambia la columna especificada en el archivo de Excel a mayúsculas
        :param N: int, El número de serie de la columna que se desea cambiar
        :param save_file_name: str, nombre del archivo fuente
        :return:(int, str), El primero es el valor de retorno de write_excel, mientras que el segundo es el nombre del archivo guardado de los datos procesados
        >>> processor = ExcelProcessor()
        >>> success, output_file = processor.process_excel_data(1, 'test_data.xlsx')
        """
        import openpyxl
        import os
        
        try:
            # Cargar el archivo Excel
            workbook = openpyxl.load_workbook(save_file_name)
            sheet = workbook.active
            
            # Convertir la columna N a mayúsculas
            # N es 1-indexed (la primera columna es 1)
            for row in range(1, sheet.max_row + 1):
                cell = sheet.cell(row=row, column=N)
                if cell.value is not None and isinstance(cell.value, str):
                    cell.value = cell.value.upper()
            
            # Generar nombre del archivo de salida
            base_name = os.path.splitext(save_file_name)[0]
            extension = os.path.splitext(save_file_name)[1]
            output_file_name = f"{base_name}_processed{extension}"
            
            # Guardar el archivo procesado
            workbook.save(output_file_name)
            workbook.close()
            
            # Llamar a write_excel (asumiendo que existe en la clase)
            result = self.write_excel(output_file_name)
            
            return (result, output_file_name)
            
        except Exception as e:
            # En caso de error, retornar código de error
            return (0, "")