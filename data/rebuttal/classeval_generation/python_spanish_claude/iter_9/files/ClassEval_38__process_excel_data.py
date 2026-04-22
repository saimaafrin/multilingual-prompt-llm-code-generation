class _M:
    def process_excel_data(self, N, save_file_name):
        """
        Cambia la columna especificada en el archivo de Excel a mayúsculas
        :param N: int, El número de serie de la columna que se desea cambiar
        :param save_file_name: str, nombre del archivo fuente
        :return:(int, str), El primero es el valor de retorno de write_excel, mientras que el segundo es el nombre del archivo guardado de los datos procesados
        >>> processor = ExcelProcessor()
        >>> success, output_file = processor.process_excel_data(1, 'test_data.xlsx')
        """
        import openpyxl
        
        # Read the Excel file
        try:
            workbook = openpyxl.load_workbook(save_file_name)
            sheet = workbook.active
            
            # Convert the specified column (N) to uppercase
            # N is 1-indexed (1 = column A, 2 = column B, etc.)
            for row in range(1, sheet.max_row + 1):
                cell = sheet.cell(row=row, column=N)
                if cell.value is not None and isinstance(cell.value, str):
                    cell.value = cell.value.upper()
            
            # Generate output filename
            output_file_name = save_file_name.replace('.xlsx', '_processed.xlsx')
            
            # Save the modified workbook
            workbook.save(output_file_name)
            workbook.close()
            
            # Assuming write_excel returns 1 for success
            return (1, output_file_name)
            
        except Exception as e:
            # Return 0 for failure
            return (0, str(e))